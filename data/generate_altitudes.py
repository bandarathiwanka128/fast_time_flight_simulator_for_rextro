import openpyxl
import random
from openpyxl.utils import column_index_from_string

# Specify the input file and column containing the dataset
input_file = 'Flight_Info.xlsx'
column_path = 'N'  # Update with the appropriate column letter

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define the headers for the columns
header = ['Column Path', 'Generated Data']
sheet.append(header)

# Open the input file
input_workbook = openpyxl.load_workbook(input_file)
input_sheet = input_workbook.active

# Iterate over the rows in the input file
for row in input_sheet.iter_rows(values_only=True):
    column_data = str(row[column_index_from_string(column_path) - 1])  # Get the data from the specified column

    # Get the comma-separated values from the row
    comma_values = column_data.split(',')
    num_values = len(comma_values)

    # Generate the values based on the specified conditions
    generated_values = []
    #constant_value = random.choice([450, 425, 400, 475, 500, 490])
    constant_value = random.choice([36000, 33000, 38000, 41000, 31000])


    for index, value in enumerate(comma_values):
        if index == 0 or index == (num_values - 1):
            generated_values.append(random.choice([12000, 7000]))
            #generated_values.append(0)
        else:
            generated_values.append(constant_value)

    #generated_values.pop()

    # Format the generated values with square brackets and append the row to the sheet
    generated_data = '[' + ','.join(str(val) for val in generated_values) + ']'
    row_data = [column_data, generated_data]
    sheet.append(row_data)

# Save the workbook as an Excel file
workbook.save('2.xlsx')
