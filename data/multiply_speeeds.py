import openpyxl
from openpyxl.utils import get_column_letter
import ast

# Function to multiply the numbers in a list by 8 and convert it back to the string format
def multiply_and_format(lst):
    multiplied_list = [str(num * 8) for num in lst]
    return '[' + ', '.join(multiplied_list) + ']'

# Specify the input file and column containing the dataset
input_file = 'Flight_Info.xlsx'
column_path = 'Speed'  # Update with the appropriate column letter

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define the headers for the columns
header = ['Column Path', 'Generated Data']
sheet.append(header)

# Open the input file
input_workbook = openpyxl.load_workbook(input_file)
input_sheet = input_workbook.active

# Find the index of the column containing the dataset
column_index = None
for cell in input_sheet[1]:  # Assuming the column names are in the first row
    print(cell.value);
    if cell.value == column_path:
        column_index = cell.column
        break

if column_index is None:
    print(f"Column '{column_path}' not found in the input sheet.")
    exit()

# Process the data and write the results to the new workbook
for row in input_sheet.iter_rows(values_only=True):
    data_str = row[column_index - 1]  # Adjust for 0-indexed column
    try:
        data_list = ast.literal_eval(data_str)
        multiplied_data = multiply_and_format(data_list)
        sheet.append([data_str, multiplied_data])
    except (SyntaxError, ValueError):
        # Error occurred while evaluating the string as a list, skipping this row
        pass

# Save the new workbook to a file
output_file = 'output.xlsx'
workbook.save(output_file)

print("Data successfully multiplied and saved to 'output.xlsx'.")
